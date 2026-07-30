# -*- coding: utf-8 -*-
"""
xlsx_reader.py
从 xlsx 文件中提取 PLC 项目配置信息，并将其转换为结构化 Python 对象。

支持的 xlsx 格式（基于 PLC程序完整.xlsx）：
─────────────────────────────────────────────────────────────
区段 1  设备清单（从首行"名称/型号"标题开始）
  行格式：名称 | 型号 | ... | 数量
  示例：cpu | SIMATIC s7-1200 1214C ... | 1

区段 2  I/O 点表（以"xxx点表"或"数字输入量模块N"为列组标题开始）
  每 4 列为一个模块组：名称 | 类型 | 地址 | （分隔列）
  示例：1#小火反馈 | Bool | %I0.0

区段 3  DB 块定义（以功能名称行开始）
  包含：DB 块名称行、列表头行、变量行（信号数据/数字量信号/内部数据/通讯数据）
  以及功能描述行
─────────────────────────────────────────────────────────────
使用方法：
  from scdw.xlsx.reader import read_plc_project_xlsx
  project = read_plc_project_xlsx("data/xlsx/PLC程序完整.xlsx")
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
except ImportError as exc:
    raise ImportError("请先安装 openpyxl：pip install openpyxl") from exc


# ── 数据模型 ──────────────────────────────────────────────────────────────────
@dataclass
class HardwareDevice:
    """硬件设备条目。"""

    category: str        # 如 "cpu"、"通讯模块"、"数字输入量模块"
    model_full: str      # 完整型号字符串，如 "SIMATIC s7-1200 1214C ... 6ES7 214-1BG40-0XB0"
    quantity: int = 1

    @property
    def order_number(self) -> Optional[str]:
        """从型号字符串中提取订货号（如 6ES7 214-1BG40-0XB0）。"""
        # 修正match方式，支持末尾带有版本号：6ES7 214-1BG40-0XB0/V4.5
        match = re.search(r"6[A-Z]{2}\d[\w\s\-]+(?:/V\d+\.\d+)?", self.model_full)
        # match = re.search(r"6[A-Z]{2}\d[\w\s\-]+", self.model_full)
        return match.group(0).strip() if match else None


@dataclass
class IOTag:
    """I/O 变量条目。"""

    name: str
    data_type: str
    address: str
    module_group: str = ""   # 所属模块组名称，如 "cpu点表"、"数字输入量模块1"
    comment: str = ""


@dataclass
class DBVariableEntry:
    """DB 块中的变量条目。"""

    name: str
    data_type: str
    offset: Any              # 偏移量（float / int / str）
    group: str = ""          # 所属变量组
    comment: str = ""


@dataclass
class DBBlockSpec:
    """全局 DB 块规格。"""

    function_name: str           # 功能名称，如 "风机输出功能"
    variables: List[DBVariableEntry] = field(default_factory=list)
    description: str = ""        # 功能描述文本


@dataclass
class FunctionLogicSpec:
    """
    程序功能逻辑规格：每条对应 xlsx 中一个功能区块的名称与逻辑描述。
    描述文本作为 LAD 整体规划的需求输入，不预设块或Network边界。
    关联 DB 块名称记录在 db_block_name，供同一对话中的LLM规划符号引用。
    """

    function_name: str       # 功能名称，如 "风机输出功能"
    description: str         # 逻辑需求文本，由整体规划决定块和Network组织
    db_block_name: str = "" # 关联的 DB 块名称（与 DBBlockSpec.function_name 一致）
    block_index: int = 0     # 建议的 LAD 块编号（从 200 起，避免与 DB 编号 100+ 冲突）


@dataclass
class PLCProjectSpec:
    """完整的 PLC 项目规格，由 xlsx 解析得到。"""

    hardware: List[HardwareDevice] = field(default_factory=list)
    io_tags: List[IOTag] = field(default_factory=list)
    db_blocks: List[DBBlockSpec] = field(default_factory=list)
    logic_functions: List[FunctionLogicSpec] = field(default_factory=list)  # 各功能的 LAD 逻辑描述


# ── 内部辅助 ──────────────────────────────────────────────────────────────────
def _cell_str(cell_value: Any) -> str:
    """将单元格值转为去除空白的字符串。"""
    if cell_value is None:
        return ""
    return str(cell_value).strip()


def _row_values(row) -> List[Any]:
    """将 openpyxl Row 对象转为单元格值列表。"""
    return [cell.value for cell in row]


def _is_mostly_empty(values: List[Any], min_non_none: int = 1) -> bool:
    """若行中非 None 的单元格数量少于 min_non_none，认为该行"基本为空"。"""
    return sum(1 for v in values if v is not None) < min_non_none


# ── 区段 1：设备清单解析 ──────────────────────────────────────────────────────
def _parse_hardware(rows: List[List[Any]]) -> List[HardwareDevice]:
    """
    从工作表行列表中提取设备清单。
    识别规则：首行含"名称"与"型号"关键字，后续行每行包含设备类别和型号。
    遇到连续两个空行或新区段标志时停止。
    """
    devices: List[HardwareDevice] = []
    in_section = False
    empty_count = 0

    for values in rows:
        row_text = [_cell_str(v) for v in values]
        first_two = row_text[:2]

        # 识别表头行
        if not in_section:
            if "名称" in first_two and "型号" in first_two:
                in_section = True
            continue

        if _is_mostly_empty(values):
            empty_count += 1
            if empty_count >= 2:
                break
            continue
        else:
            empty_count = 0

        category = row_text[0]
        model = row_text[1]
        if not category or not model:
            continue

        # 数量在列索引 7（基于既有格式）
        qty_raw = values[7] if len(values) > 7 else None
        try:
            qty = int(qty_raw) if qty_raw is not None else 1
        except (ValueError, TypeError):
            qty = 1

        devices.append(HardwareDevice(category=category, model_full=model, quantity=qty))

    return devices


# ── 区段 2：I/O 点表解析 ──────────────────────────────────────────────────────
_MODULE_HEADER_RE = re.compile(
    # 要求"数字输入量模块"后必须跟数字，以区别于硬件清单中的类别名称
    r"(cpu点表|数字输入量模块\d+|输入模块\d+|输出模块\d+|io模块\d+)",
    re.IGNORECASE,
)
_COL_HEADER_KEYWORDS = {"名称", "类型", "地址"}


def _is_module_header_row(values: List[Any]) -> bool:
    """检查该行是否为点表区段的模块分组标题行。"""
    for v in values:
        if v and _MODULE_HEADER_RE.search(_cell_str(v)):
            return True
    return False


def _is_column_header_row(values: List[Any]) -> bool:
    """检查该行是否为点表的列名行（含名称/类型/地址关键字）。"""
    text_set = {_cell_str(v) for v in values if v is not None}
    return len(_COL_HEADER_KEYWORDS & text_set) >= 2


def _parse_module_groups(header_values: List[Any]) -> Dict[int, str]:
    """
    从模块标题行提取 {列索引: 模块名称} 映射。
    每组中第一个非空列即为该模块名，后续非空列标记同一组模块。
    """
    groups: Dict[int, str] = {}
    current_group = ""
    for col_idx, val in enumerate(header_values):
        text = _cell_str(val)
        if text:
            current_group = text
        if current_group:
            groups[col_idx] = current_group
    return groups


def _parse_io_tags(rows: List[List[Any]]) -> List[IOTag]:
    """
    从工作表行列表中提取所有 I/O 点表变量。
    支持多模块横向排列（每 4 列一组：名称/类型/地址/分隔）。
    """
    tags: List[IOTag] = []
    in_section = False
    col_to_group: Dict[int, str] = {}
    consecutive_empty = 0

    for values in rows:
        # 识别模块标题行 → 提取列-模块映射
        if not in_section or _is_module_header_row(values):
            if _is_module_header_row(values):
                col_to_group = _parse_module_groups(values)
                in_section = True
                consecutive_empty = 0
            continue

        # 识别列名行 → 跳过
        if _is_column_header_row(values):
            consecutive_empty = 0
            continue

        if _is_mostly_empty(values):
            consecutive_empty += 1
            if consecutive_empty >= 3:
                # 连续空行 3 行认为点表结束
                break
            continue
        else:
            consecutive_empty = 0

        # 每4列为一个模块组：名称(0)/类型(1)/地址(2)/separator(3)
        num_cols = len(values)
        for group_start in range(0, num_cols, 4):
            name_val = values[group_start] if group_start < num_cols else None
            type_val = values[group_start + 1] if group_start + 1 < num_cols else None
            addr_val = values[group_start + 2] if group_start + 2 < num_cols else None

            name = _cell_str(name_val)
            dtype = _cell_str(type_val)
            addr = _cell_str(addr_val)

            if not name or not addr:
                continue

            module_name = col_to_group.get(group_start, "")

            tags.append(
                IOTag(
                    name=name,
                    data_type=dtype or "Bool",
                    address=addr,
                    module_group=module_name,
                )
            )

    return tags


# ── 区段 3：DB 块解析 ─────────────────────────────────────────────────────────
_DB_SECTION_KEYWORDS = {"DB块名称", "信号数据", "数字量信号", "内部数据", "通讯数据"}
_DESCRIPTION_KEYWORD = "功能描述"


def _is_db_col_group_header(values: List[Any]) -> bool:
    """检查是否为 DB 变量列组标题行（含"信号数据"等关键字）。"""
    texts = {_cell_str(v) for v in values if v is not None}
    return len(texts & _DB_SECTION_KEYWORDS) >= 2


def _is_db_sub_col_header(values: List[Any]) -> bool:
    """检查是否为 DB 变量子列名行（含"名称"/"类型"/"偏移量"关键字）。"""
    texts = {_cell_str(v) for v in values if v is not None}
    return "偏移量" in texts and "名称" in texts


def _parse_db_blocks(rows: List[List[Any]]) -> List[DBBlockSpec]:
    """
    两阶段解析 DB 块定义。

    阶段 1：扫描所有"列组标题行"（含"信号数据"/"数字量信号"等关键字），
            以此确定每个 DB 块的精确边界，不依赖行间距假设。

    阶段 2：对每个 DB 块范围独立解析：
            - 向前查找功能名称行（列组标题行之前的最后一个非空、非标题行）
            - 跳过子列名行（含"偏移量"关键字）
            - 向后查找"功能描述"行（任意列包含该关键字即可）
            - 按 4 列步长提取变量（列 1/2/3，列 5/6/7，…），组名从列组标题的
              col+1 起步（col 0 是标签"DB块名称"，不作为组名）

    修复的已知问题：
      - 组名偏移一位（group_col_map[0] = "DB块名称" 被误用）
      - 多 DB 块时内层循环无法正确终止
      - "功能描述"检测仅看 col 0，合并单元格时失效
    """
    total = len(rows)

    # ── 阶段 1：找到所有列组标题行的索引 ────────────────────────────────────
    header_indices: List[int] = [
        i for i in range(total) if _is_db_col_group_header(rows[i])
    ]
    if not header_indices:
        return []

    db_blocks: List[DBBlockSpec] = []

    for seq, hdr_i in enumerate(header_indices):

        # ── 向前查找功能名称行（标题行之前最近的非空行，col 0 非空）──────────
        function_name = ""
        for back in range(hdr_i - 1, max(hdr_i - 8, -1), -1):
            bvals = rows[back]
            val0 = _cell_str(bvals[0]) if bvals else ""
            if val0 and not _is_db_col_group_header(bvals):
                function_name = val0
                break

        # ── 本块内容截止位置：下一个列组标题行（或文件末尾）──────────────────
        next_hdr = header_indices[seq + 1] if seq + 1 < len(header_indices) else total

        # ── 构建"列索引 → 组名"映射，跳过 col 0 标签列 ────────────────────────
        group_col_map: Dict[int, str] = {}
        current_group = ""
        for col_idx, val in enumerate(rows[hdr_i]):
            if col_idx == 0:
                continue  # col 0 是 "DB块名称" 标签，不作为组名
            text = _cell_str(val)
            if text:
                current_group = text
            if current_group:
                group_col_map[col_idx] = current_group

        # ── 跳过子列名行（紧随列组标题行之后，含"偏移量"关键字）──────────────
        data_start = hdr_i + 1
        if data_start < total and _is_db_sub_col_header(rows[data_start]):
            data_start += 1

        # ── 查找"功能描述"行：任意单元格以关键字开头即算 ──────────────────────
        desc_i: Optional[int] = None
        for j in range(data_start, next_hdr):
            for v in rows[j]:
                if _cell_str(v).startswith(_DESCRIPTION_KEYWORD):
                    desc_i = j
                    break
            if desc_i is not None:
                break

        # ── 解析变量数据行（data_start → desc_i 或 next_hdr）────────────────
        var_end = desc_i if desc_i is not None else next_hdr
        db_block = DBBlockSpec(function_name=function_name)

        for row in rows[data_start:var_end]:
            # 跳过空行及重复出现的子列名行
            if _is_mostly_empty(row, min_non_none=2) or _is_db_sub_col_header(row):
                continue

            num_cols = len(row)
            for group_start in range(0, num_cols, 4):
                # 列布局：col 0=标签/空, col 1=名称, col 2=类型, col 3=偏移量
                # 下一组: col 4=空, col 5=名称, col 6=类型, col 7=偏移量 …
                name_v = row[group_start + 1] if group_start + 1 < num_cols else None
                type_v = row[group_start + 2] if group_start + 2 < num_cols else None
                offset_v = row[group_start + 3] if group_start + 3 < num_cols else None

                name = _cell_str(name_v)
                dtype = _cell_str(type_v)
                if not name or not dtype:
                    continue

                # 组名从 group_start+1 列查找（跳过 col 0 标签偏移）
                group_name = group_col_map.get(group_start + 1, "")
                db_block.variables.append(
                    DBVariableEntry(
                        name=name,
                        data_type=dtype,
                        offset=offset_v,
                        group=group_name,
                    )
                )

        # ── 解析功能描述 ──────────────────────────────────────────────────────
        if desc_i is not None:
            desc_row = rows[desc_i]
            desc_parts: List[str] = []
            found_kw = False
            for v in desc_row:
                s = _cell_str(v)
                if not found_kw:
                    if s.startswith(_DESCRIPTION_KEYWORD):
                        found_kw = True
                        # 关键字之后可能紧跟描述文本（如"功能描述：xxx"）
                        after = s[len(_DESCRIPTION_KEYWORD):].lstrip("：: \t")
                        if after:
                            desc_parts.append(after)
                elif s:
                    desc_parts.append(s)
            db_block.description = " ".join(desc_parts)

            # 若同行无描述文本，尝试下一行
            if not db_block.description and desc_i + 1 < total:
                next_row = rows[desc_i + 1]
                db_block.description = " ".join(
                    _cell_str(v) for v in next_row if v is not None
                ).strip()

        db_blocks.append(db_block)

    return db_blocks


def _build_logic_functions(db_blocks: "List[DBBlockSpec]") -> "List[FunctionLogicSpec]":
    """
    从已解析的 DB 块列表中提取程序功能逻辑规格。
    每个含描述的 DB 块生成一条 FunctionLogicSpec。
    block_index 从 200 开始，避免与 DB 编号（100+）冲突。
    """
    result: List[FunctionLogicSpec] = []
    for i, db in enumerate(db_blocks):
        if db.description:
            result.append(
                FunctionLogicSpec(
                    function_name=db.function_name,
                    description=db.description,
                    db_block_name=db.function_name,
                    block_index=200 + i,
                )
            )
    return result


# ── 公共入口 ──────────────────────────────────────────────────────────────────
def read_plc_project_xlsx(filepath: str, sheet_name: str = "Sheet1") -> PLCProjectSpec:
    """
    读取 xlsx 文件并解析为 PLCProjectSpec 对象。

    Args:
        filepath: xlsx 文件路径
        sheet_name: 要解析的工作表名称（默认 "Sheet1"）

    Returns:
        PLCProjectSpec 包含：hardware（设备清单）、io_tags（I/O 点表）、db_blocks（DB 块）

    Raises:
        FileNotFoundError: 文件不存在
        KeyError: 工作表不存在
    """
    import os

    if not os.path.isfile(filepath):
        raise FileNotFoundError(f"xlsx 文件不存在：{filepath}")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"工作表 '{sheet_name}' 不存在，可用：{wb.sheetnames}")

    ws = wb[sheet_name]
    all_rows: List[List[Any]] = [_row_values(row) for row in ws.iter_rows()]
    wb.close()

    # ── 识别各区段的行范围 ──────────────────────────────────────────────────
    hardware_start = _find_hardware_section_start(all_rows)
    io_start = _find_io_section_start(all_rows)
    db_start = _find_db_section_start(all_rows)

    hardware = _parse_hardware(all_rows[hardware_start:io_start] if io_start else all_rows[hardware_start:])
    io_tags = _parse_io_tags(all_rows[io_start:db_start] if db_start else all_rows[io_start:]) if io_start is not None else []
    db_blocks = _parse_db_blocks(all_rows[db_start:]) if db_start is not None else []
    logic_functions = _build_logic_functions(db_blocks)

    return PLCProjectSpec(hardware=hardware, io_tags=io_tags, db_blocks=db_blocks, logic_functions=logic_functions)


def _find_hardware_section_start(rows: List[List[Any]]) -> int:
    """查找设备清单区段的起始行索引（含标题行）。"""
    for idx, values in enumerate(rows):
        texts = [_cell_str(v) for v in values if v is not None]
        if "名称" in texts and "型号" in texts:
            return idx
    return 0


def _find_io_section_start(rows: List[List[Any]]) -> Optional[int]:
    """查找 I/O 点表区段的起始行索引。"""
    for idx, values in enumerate(rows):
        if _is_module_header_row(values):
            return idx
    return None


def _find_db_section_start(rows: List[List[Any]]) -> Optional[int]:
    """查找 DB 块区段的起始行索引（首个 DB 列组标题行的前一行）。"""
    for idx, values in enumerate(rows):
        if _is_db_col_group_header(values):
            # DB 区段从功能名称行开始（DB 标题行的前一个非空行）
            for back in range(idx - 1, max(idx - 6, -1), -1):
                back_vals = rows[back]
                if not _is_mostly_empty(back_vals) and _cell_str(back_vals[0]):
                    return back
            return idx
    return None


# ── 调试辅助：打印解析结果摘要 ───────────────────────────────────────────────
def print_project_summary(spec: PLCProjectSpec) -> None:
    """打印 PLCProjectSpec 的概要信息，方便调试。"""
    print(f"\n{'='*50}")
    print(f"设备清单（{len(spec.hardware)} 项）")
    for hw in spec.hardware:
        print(f"  [{hw.category}] {hw.model_full[:60]}  ×{hw.quantity}  订货号:{hw.order_number}")

    print(f"\nI/O 变量表（{len(spec.io_tags)} 个变量）")
    groups: Dict[str, int] = {}
    for tag in spec.io_tags:
        groups[tag.module_group] = groups.get(tag.module_group, 0) + 1
    for grp, cnt in groups.items():
        print(f"  [{grp}] {cnt} 个变量")

    print(f"\nDB 块（{len(spec.db_blocks)} 个）")
    for db in spec.db_blocks:
        print(f"  [{db.function_name}] {len(db.variables)} 个变量  描述长度:{len(db.description)} 字符")
    print("=" * 50)
