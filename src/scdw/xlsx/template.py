# -*- coding: utf-8 -*-
"""
create_template.py
生成与 xlsx_reader.py 兼容的 PLC 项目 xlsx 模板文件。

运行：
    python data/create_template.py
输出：
    data/PLC程序模板.xlsx

─────────────────────────────────────────────────────────────
Sheet1 结构
─────────────────────────────────────────────────────────────
区段 1  设备清单
  headerrow  : A=名称  B=型号  C~G留空  H=数量
  datarows   : A=类别  B=完整型号  H=数量

区段 2  I/O 点表（紧跟设备清单后，两个空行隔开）
  modulerow  : A=cpu点表  E=数字输入量模块1  I=数字输入量模块2
  colhdrrow  : A=名称 B=类型 C=地址 D=空  E=名称 F=类型 G=地址 ...
  datarows   : 每4列一组  名称|类型|地址|空

区段 3  DB 块（紧跟 I/O 点表后，三个空行隔开）
  funcrow    : A=功能名称
  grphdrrow  : A=DB块名称  B=信号数据  F=数字量信号  J=内部数据  N=通讯数据
  colhdrrow  : B=名称 C=类型 D=偏移量  F=名称 G=类型 H=偏移量 ...
  datarows   : 每4列一组（col0空/标签, col1=名称, col2=类型, col3=偏移量）
  descrow    : A=功能描述  B=描述正文
─────────────────────────────────────────────────────────────
"""
from __future__ import annotations

import os
from typing import Any, List, Optional, Tuple
from scdw.common.paths import XLSX_DATA_DIR

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise ImportError("请先安装 openpyxl：pip install openpyxl") from exc


# ── 样式常量 ──────────────────────────────────────────────────────────────────
_BLUE_FILL   = PatternFill("solid", fgColor="4472C4")
_GREEN_FILL  = PatternFill("solid", fgColor="70AD47")
_YELLOW_FILL = PatternFill("solid", fgColor="FFD966")
_GRAY_FILL   = PatternFill("solid", fgColor="D9D9D9")
_WHITE_FONT  = Font(bold=True, color="FFFFFF")
_BOLD_FONT   = Font(bold=True)
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _set(ws, row: int, col: int, value: Any,
         fill=None, font=None, alignment=None, border=None) -> None:
    """写入单元格并应用样式。"""
    cell = ws.cell(row=row, column=col, value=value)
    if fill:      cell.fill      = fill
    if font:      cell.font      = font
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border


def _hdr(ws, row: int, col: int, value: str, fill=None) -> None:
    _set(ws, row, col, value,
         fill=fill or _GRAY_FILL, font=_BOLD_FONT,
         alignment=_CENTER, border=_THIN_BORDER)


def _data(ws, row: int, col: int, value: Any) -> None:
    _set(ws, row, col, value, alignment=_LEFT, border=_THIN_BORDER)


# ── 区段 1：设备清单 ──────────────────────────────────────────────────────────
def _write_hardware(ws, start_row: int) -> int:
    """写入设备清单，返回写完后的下一个可用行号。"""
    # 合并标题区
    ws.merge_cells(f"A{start_row}:H{start_row}")
    _set(ws, start_row, 1, "设备清单",
         fill=_BLUE_FILL, font=Font(bold=True, color="FFFFFF", size=13),
         alignment=_CENTER)
    r = start_row + 1

    # 列标题行：A=名称 B=型号 C-G留空 H=数量
    for col, label in [(1, "名称"), (2, "型号"), (8, "数量")]:
        _hdr(ws, r, col, label, fill=_BLUE_FILL)
        if col not in (1, 2, 8):
            _hdr(ws, r, col, "")
    # C~G 表头留空（保持边框）
    for col in range(3, 8):
        _hdr(ws, r, col, "")
    r += 1

    # 示例设备数据（A=类别, B=完整型号, H=数量）
    devices = [
        ("cpu",         "SIMATIC S7-1200 1214C AC/DC/Rly 6ES7 214-1BG40-0XB0", 1),
        ("通讯模块",    "CM 1241 (RS422/485) 6ES7 241-1CH32-0XB0",              1),
        ("数字输入量模块", "DI 16x24VDC/DQ 16xRelay 6ES7 223-1PL30-0XB0",      2),
    ]
    for category, model, qty in devices:
        _data(ws, r, 1, category)
        _data(ws, r, 2, model)
        for col in range(3, 8):
            _data(ws, r, col, None)
        _data(ws, r, 8, qty)
        r += 1

    return r  # 结束后的下一行（调用方负责插入空行）


# ── 区段 2：I/O 点表 ──────────────────────────────────────────────────────────
def _write_io_section(ws, start_row: int) -> int:
    """
    写入 I/O 点表区段。
    布局：每组占 4 列（名称/类型/地址/空），横向排列多模块。
    返回写完后的下一个可用行号。
    """
    # ── 模块组定义 ────────────────────────────────────────────────────────────
    # (模块名, [(变量名, 数据类型, 地址), ...])
    io_groups: List[Tuple[str, List[Tuple[str, str, str]]]] = [
        ("cpu点表", [
            ("启动按钮",   "Bool", "%I0.0"),
            ("停止按钮",   "Bool", "%I0.1"),
            ("急停按钮",   "Bool", "%I0.2"),
            ("复位按钮",   "Bool", "%I0.3"),
            ("手动/自动",  "Bool", "%I0.4"),
            ("风机运行反馈","Bool", "%I0.5"),
            ("风机故障",   "Bool", "%I0.6"),
            ("风压报警",   "Bool", "%I0.7"),
            ("蜂鸣器",     "Bool", "%Q0.0"),
            ("指示灯-运行","Bool", "%Q0.1"),
            ("指示灯-故障","Bool", "%Q0.2"),
        ]),
        ("数字输入量模块1", [
            ("1#小火反馈",  "Bool", "%I2.0"),
            ("1#大火反馈",  "Bool", "%I2.1"),
            ("1#故障反馈",  "Bool", "%I2.2"),
            ("2#小火反馈",  "Bool", "%I2.3"),
            ("2#大火反馈",  "Bool", "%I2.4"),
            ("2#故障反馈",  "Bool", "%I2.5"),
            ("1#小火输出",  "Bool", "%Q1.0"),
            ("1#大火输出",  "Bool", "%Q1.1"),
            ("1#复位输出",  "Bool", "%Q1.2"),
            ("2#小火输出",  "Bool", "%Q1.3"),
            ("2#大火输出",  "Bool", "%Q1.4"),
            ("2#复位输出",  "Bool", "%Q1.5"),
        ]),
        ("数字输入量模块2", [
            ("3#小火反馈",  "Bool", "%I4.0"),
            ("3#大火反馈",  "Bool", "%I4.1"),
            ("3#故障反馈",  "Bool", "%I4.2"),
            ("4#小火反馈",  "Bool", "%I4.3"),
            ("4#大火反馈",  "Bool", "%I4.4"),
            ("4#故障反馈",  "Bool", "%I4.5"),
            ("3#小火输出",  "Bool", "%Q2.0"),
            ("3#大火输出",  "Bool", "%Q2.1"),
            ("3#复位输出",  "Bool", "%Q2.2"),
            ("4#小火输出",  "Bool", "%Q2.3"),
            ("4#大火输出",  "Bool", "%Q2.4"),
            ("4#复位输出",  "Bool", "%Q2.5"),
        ]),
    ]

    num_groups = len(io_groups)
    # 每组占 4 列（A-D / E-H / I-L …）
    # 模块标题行
    r = start_row
    ws.merge_cells(
        f"A{r}:{get_column_letter(num_groups * 4)}{r}"
    )
    _set(ws, r, 1, "I/O 点表",
         fill=_GREEN_FILL, font=Font(bold=True, color="FFFFFF", size=13),
         alignment=_CENTER)
    r += 1

    # 模块名行
    for g_idx, (mod_name, _) in enumerate(io_groups):
        col_start = g_idx * 4 + 1
        ws.merge_cells(
            f"{get_column_letter(col_start)}{r}:{get_column_letter(col_start + 2)}{r}"
        )
        _hdr(ws, r, col_start, mod_name, fill=_GREEN_FILL)
        _set(ws, r, col_start, mod_name,
             fill=_GREEN_FILL, font=_WHITE_FONT, alignment=_CENTER)
        _data(ws, r, col_start + 3, None)  # 分隔列
    r += 1

    # 列名行（名称/类型/地址/空）× 每组
    for g_idx in range(num_groups):
        col_start = g_idx * 4 + 1
        for offset, label in enumerate(["名称", "类型", "地址", ""]):
            _hdr(ws, r, col_start + offset, label, fill=_GRAY_FILL)
    r += 1

    # 数据行（纵向对齐各组）
    max_rows = max(len(tags) for _, tags in io_groups)
    for entry_idx in range(max_rows):
        for g_idx, (_, tags) in enumerate(io_groups):
            col_start = g_idx * 4 + 1
            if entry_idx < len(tags):
                name, dtype, addr = tags[entry_idx]
                _data(ws, r, col_start,     name)
                _data(ws, r, col_start + 1, dtype)
                _data(ws, r, col_start + 2, addr)
            _data(ws, r, col_start + 3, None)  # 分隔列
        r += 1

    return r


# ── 区段 3：DB 块 ─────────────────────────────────────────────────────────────
def _write_db_block(
    ws,
    start_row: int,
    function_name: str,
    signal_vars: List[Tuple[str, str, Any]],       # 信号数据 (名称, 类型, 偏移量)
    digital_vars: List[Tuple[str, str, Any]],      # 数字量信号
    internal_vars: List[Tuple[str, str, Any]],     # 内部数据
    comm_vars: List[Tuple[str, str, Any]],         # 通讯数据
    description: str,
) -> int:
    """
    写入一个 DB 块。列布局（每组4列，col0=标签/空）：
      col A : DB块名称标签 / 变量行为空
      col B : 信号数据标签 / 名称
      col C : 类型
      col D : 偏移量
      col E : 空（分隔）
      col F : 数字量信号标签 / 名称
      col G : 类型
      col H : 偏移量
      col I : 空（分隔）
      col J : 内部数据标签 / 名称
      col K : 类型
      col L : 偏移量
      col M : 空（分隔）
      col N : 通讯数据标签 / 名称
      col O : 类型
      col P : 偏移量
    返回下一个可用行号。
    """

    # ── 功能名称行 ────────────────────────────────────────────────────────────
    r = start_row
    ws.merge_cells(f"A{r}:P{r}")
    _set(ws, r, 1, function_name,
         fill=_YELLOW_FILL, font=Font(bold=True, size=12),
         alignment=_CENTER)
    r += 1

    # ── 列组标题行 ────────────────────────────────────────────────────────────
    # A=DB块名称, B=信号数据(占4列B-E), F=数字量信号(F-I), J=内部数据(J-M), N=通讯数据(N-P)
    group_headers = [
        (1, "DB块名称", _GRAY_FILL),
        (2, "信号数据",   _BLUE_FILL),
        (6, "数字量信号", _GREEN_FILL),
        (10, "内部数据",  _YELLOW_FILL),
        (14, "通讯数据",  PatternFill("solid", fgColor="ED7D31")),
    ]
    for col, label, fill in group_headers:
        _set(ws, r, col, label,
             fill=fill, font=_WHITE_FONT, alignment=_CENTER, border=_THIN_BORDER)
    # 填充空列
    for col in [3, 4, 5, 7, 8, 9, 11, 12, 13, 15, 16]:
        _hdr(ws, r, col, "")
    r += 1

    # ── 子列名行 ────────────────────────────────────────────────────────────
    sub_cols = [
        # 每组: 空(A/E/I/M), 名称, 类型, 偏移量
        (1, ""),    (2, "名称"),   (3, "类型"),   (4, "偏移量"),
        (5, ""),    (6, "名称"),   (7, "类型"),   (8, "偏移量"),
        (9, ""),    (10, "名称"),  (11, "类型"),  (12, "偏移量"),
        (13, ""),   (14, "名称"),  (15, "类型"),  (16, "偏移量"),
    ]
    for col, label in sub_cols:
        _hdr(ws, r, col, label)
    r += 1

    # ── 变量数据行 ──────────────────────────────────────────────────────────
    all_groups = [signal_vars, digital_vars, internal_vars, comm_vars]
    max_rows = max(len(g) for g in all_groups)
    for entry_idx in range(max_rows):
        _data(ws, r, 1, None)  # col A 空
        for g_idx, group in enumerate(all_groups):
            base = g_idx * 4 + 2   # B=2, F=6, J=10, N=14
            if entry_idx < len(group):
                name, dtype, offset = group[entry_idx]
                _data(ws, r, base,     name)
                _data(ws, r, base + 1, dtype)
                _data(ws, r, base + 2, offset)
            sep_col = base + 3  # E=5, I=9, M=13, Q=17 — 但 Q 超出范围，跳过
            if sep_col <= 16:
                _data(ws, r, sep_col, None)
        r += 1

    # ── 功能描述行 ──────────────────────────────────────────────────────────
    _set(ws, r, 1, "功能描述",
         fill=_GRAY_FILL, font=_BOLD_FONT, alignment=_CENTER, border=_THIN_BORDER)
    desc_cell = ws.cell(row=r, column=2, value=description)
    desc_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    desc_cell.border = _THIN_BORDER
    ws.merge_cells(f"B{r}:P{r}")
    ws.row_dimensions[r].height = max(60, len(description) // 4)
    r += 1

    return r


# ── 列宽设置 ──────────────────────────────────────────────────────────────────
def _set_column_widths(ws) -> None:
    widths = {
        "A": 16, "B": 20, "C": 18, "D": 10,
        "E": 4,  "F": 20, "G": 18, "H": 10,
        "I": 4,  "J": 24, "K": 22, "L": 10,
        "M": 4,  "N": 20, "O": 18, "P": 10,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ── 主函数 ───────────────────────────────────────────────────────────────────
def create_template(output_path: str | None = None) -> None:
    """生成 XLSX 样例；未传路径时写入统一数据目录。"""
    if output_path is None:
        XLSX_DATA_DIR.mkdir(parents=True, exist_ok=True)
        output_path = str(XLSX_DATA_DIR / "PLC程序模板.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.freeze_panes = "A2"

    r = 1

    # ── 区段 1：设备清单 ──────────────────────────────────────────────────────
    r = _write_hardware(ws, r)
    r += 2  # 两个空行

    # ── 区段 2：I/O 点表 ─────────────────────────────────────────────────────
    r = _write_io_section(ws, r)
    r += 3  # 三个空行

    # ── 区段 3：DB 块示例 1 ───────────────────────────────────────────────────
    r = _write_db_block(
        ws, r,
        function_name="风机输出功能",
        signal_vars=[
            ("风压故障",     "Bool", 129.2),
            ("风机手动",     "Bool", 129.3),
            ("风机运行按钮", "Bool", 129.4),
            ("PID控制",      "Bool", 129.5),
            ("风机自动输出", "Real", 130.0),
            ("风机手动频率", "Real", 134.0),
            ("风机控制输出", "Real", 138.0),
            ("风机反馈频率", "Real", 142.0),
        ],
        digital_vars=[
            ("风机运行标志", "Bool", 10.0),
            ("风机复位标志", "Bool", 10.1),
            ("风机运行反馈", "Bool", 10.2),
            ("风机故障反馈", "Bool", 10.3),
        ],
        internal_vars=[
            ("风机过渡",           "Int",                  74),
            ("风压转换过渡",       "Real",                 76),
            ("风机转换过渡",       "Real",                 80),
            ("风机控制输出码值",   "Real",                 84),
            ("1-8区通讯延时定时器","IEC_TIMER",            88),
            ("9-15区通讯延时定时器","IEC_TIMER",          104),
            ("变频器通讯延时定时器","IEC_TIMER",          120),
            ("风机运行数据组",     "Array[0..15] of Bool", 136),
            ("风机反馈数据组",     "Array[0..15] of Bool", 138),
        ],
        comm_vars=[
            ("变频器通讯运行状态", "Int", 32),
            ("变频器通讯反馈频率", "Int", 34),
            ("变频器通讯设定运行", "Int", 36),
            ("变频器通讯设定频率", "Int", 38),
        ],
        description=(
            "当按下风机运行按钮，且风机无故障反馈时，风机运行标志置 ON，"
            "同时系统会先判断控制模式：如果是自动模式，就直接将预设的风机自动输出作为风机控制输出；"
            "如果是手动模式，会把操作员输入的风机手动频率值放大 2 倍后作为风机控制输出。\n"
            "当风机运行按钮为 OFF，风机的控制输出频率都会立即归零。\n"
            "然后系统会把最终的风机控制输出值换算成变频器能识别的百分比频率（公式：控制输出值 ÷2×100），"
            "转换成整数后通过通讯写入变频器的频率设定地址，驱动风机按设定频率运行。"
        ),
    )
    r += 2  # 两个空行

    # ── 区段 3：DB 块示例 2 ───────────────────────────────────────────────────
    r = _write_db_block(
        ws, r,
        function_name="燃烧器控制功能",
        signal_vars=[
            ("小火允许",    "Bool", 200.0),
            ("大火允许",    "Bool", 200.1),
            ("燃烧器故障",  "Bool", 200.2),
            ("点火使能",    "Bool", 200.3),
        ],
        digital_vars=[
            ("小火运行标志", "Bool", 20.0),
            ("大火运行标志", "Bool", 20.1),
            ("燃烧器复位",   "Bool", 20.2),
        ],
        internal_vars=[
            ("小火计时器",   "IEC_TIMER", 100),
            ("大火计时器",   "IEC_TIMER", 116),
            ("点火计数器",   "Int",       132),
        ],
        comm_vars=[
            ("燃烧器状态字", "Int", 50),
            ("燃烧器控制字", "Int", 52),
        ],
        description=(
            "燃烧器控制功能负责管理点火时序与火焰状态。\n"
            "首先确认小火允许条件，启动点火程序：延时后检测小火反馈，成功后切换至大火阶段。\n"
            "若点火失败超过设定次数，则触发故障报警并锁定，需手动复位后才能重新启动。"
        ),
    )

    # ── 列宽 ─────────────────────────────────────────────────────────────────
    _set_column_widths(ws)

    wb.save(output_path)
    print(f"模板已生成：{os.path.abspath(output_path)}")


if __name__ == "__main__":
    create_template()
